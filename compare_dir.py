#!/usr/bin/env python
import sys
import os
import re
import argparse
from openpyxl import Workbook
import datetime
import hashlib

#import subprocess
# Example
# python 0.compare_dir.py -p roster -a "E:\Projects\CRDW\CCCD\rawData\June2015\UIH PLUS\Daily Panel Rosters" -b "E:\Projects\CRDW\CCCD\rawData\July2015\UIH PLUS\Daily Panel Rosters"

# python 0.compare_dir.py -p claim -a "E:\Projects\CRDW\CCCD\rawData\June2015\UIH PLUS\CCCD" -b "E:\Projects\CRDW\CCCD\rawData\July2015\UIH PLUS\CCCD"


class FILEINFO:
	''' Get file size and md5 information for a file '''
	def __init__(self, filepath):
		self.filepath	= filepath
		self.size	= self.get_file_size()
		self.md5	= self.get_md5()


	def get_file_size(self):
		fh = open(self.filepath, 'rb')
		fh.seek(0,2) # move the cursor to the end of the file
		size = fh.tell()
		fh.close()
		return size

	def get_md5(self):
		fh = open(self.filepath, 'rb')
		md5 = hashlib.md5(fh.read()).hexdigest()
		fh.close()
		return md5

	#originally, I use fciv.exe (windows) or md5sum (linux)  to get md5 checksum
	#def run_cmd(self, cmd_line):
	#	try:
	#		p = subprocess.Popen(cmd_line, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
	#		[output, errmsg] = p.communicate()
	#		if p.returncode != 0:
	#			raise OSError(p.returncode, "command %r failed: %s" % (cmd_line, errmsg))
	#		return output
	#	except Exception as ex:
	#		print ex
	#		return False
	


	

def main():
	#excle_cmp_out = r"F:\compare_dir.xlsx"
	args = get_arguments()
	MD5_TO_FILEINO_A = {}
	MD5_TO_FILEINO_B = {}


	for root, dirs, files in os.walk(args.a_dir, topdown=False):
		for filename in files:
			print os.path.join(root, filename)
			fileinfo = FILEINFO(os.path.join(root, filename))
			MD5_TO_FILEINO_A[fileinfo.md5] = fileinfo
	for root, dirs, files in os.walk(args.b_dir, topdown=False):
		for filename in files:
			print os.path.join(root, filename)
			fileinfo = FILEINFO(os.path.join(root, filename))
			MD5_TO_FILEINO_B[fileinfo.md5] = fileinfo

	A = set(MD5_TO_FILEINO_A.keys())
	B = set(MD5_TO_FILEINO_B.keys())
	print "=============="
	ONLY_IN_A	= list(A - B)
	ONLY_IN_B	= list(B - A)
	COMMON_A_B	= list(A & B) # intersection
	#UNION_A_B	= list(A | B) # union

	wb = Workbook()
	ws1 = wb.active
	ws1.title = "Files only in A"

	ROW_OFFSET = 2
	ROW_HEADER = 1
	ws1.cell(row=ROW_HEADER, column=1).value = "File (A=" + args.a_dir + ")"
	#ws1.cell(row=ROW_HEADER, column=1).value = "File"
	ws1.cell(row=ROW_HEADER, column=2).value = "MD5"
	ws1.cell(row=ROW_HEADER, column=3).value = "size (byte)"



	for i in range(len(ONLY_IN_A)):
		md5 = ONLY_IN_A[i]
		ws1.cell(row=i+ROW_OFFSET, column=1).value = MD5_TO_FILEINO_A[md5].filepath
		ws1.cell(row=i+ROW_OFFSET, column=2).value = MD5_TO_FILEINO_A[md5].md5
		ws1.cell(row=i+ROW_OFFSET, column=3).value = MD5_TO_FILEINO_A[md5].size


	ws2 = wb.create_sheet()
	ws2.title = "Files only in B"
	ws2.cell(row=ROW_HEADER, column=1).value = "File (B=" + args.b_dir + ")"
	#ws2.cell(row=ROW_HEADER, column=1).value = "File"
	ws2.cell(row=ROW_HEADER, column=2).value = "MD5"
	ws2.cell(row=ROW_HEADER, column=3).value = "size (byte)"
	for i in range(len(ONLY_IN_B)):
		md5 = ONLY_IN_B[i]
		ws2.cell(row=i+ROW_OFFSET, column=1).value = MD5_TO_FILEINO_B[md5].filepath
		ws2.cell(row=i+ROW_OFFSET, column=2).value = MD5_TO_FILEINO_B[md5].md5
		ws2.cell(row=i+ROW_OFFSET, column=3).value = MD5_TO_FILEINO_B[md5].size

	ws3 = wb.create_sheet()
	ws3.title = "Files common in A and B"
	ws3.cell(row=ROW_HEADER, column=1).value = "File (A=" + args.a_dir + ")"
	#ws3.cell(row=ROW_HEADER, column=1).value = "File"
	ws3.cell(row=ROW_HEADER, column=2).value = "MD5"
	ws3.cell(row=ROW_HEADER, column=3).value = "size (byte)"
	ws3.cell(row=ROW_HEADER, column=5).value = "File (B=" + args.b_dir + ")"
	#ws3.cell(row=ROW_HEADER, column=5).value = "File"
	ws3.cell(row=ROW_HEADER, column=6).value = "MD5"
	ws3.cell(row=ROW_HEADER, column=7).value = "size (byte)"

	for i in range(len(COMMON_A_B)):
		md5 = COMMON_A_B[i]
		ws3.cell(row=i+ROW_OFFSET, column=1).value = MD5_TO_FILEINO_A[md5].filepath
		ws3.cell(row=i+ROW_OFFSET, column=2).value = MD5_TO_FILEINO_A[md5].md5
		ws3.cell(row=i+ROW_OFFSET, column=3).value = MD5_TO_FILEINO_A[md5].size
		#ws3.cell(row=i+ROW_OFFSET, column=4).value = "'=="

		ws3.cell(row=i+ROW_OFFSET, column=5).value = MD5_TO_FILEINO_B[md5].filepath
		ws3.cell(row=i+ROW_OFFSET, column=6).value = MD5_TO_FILEINO_B[md5].md5
		ws3.cell(row=i+ROW_OFFSET, column=7).value = MD5_TO_FILEINO_B[md5].size

	wb.save(args.excel_out)
	print "The comparison results are saved in: ", args.excel_out
	

def get_arguments():
	main_description = '''\
	Compare the files in two directories using MD5 checksum.
	You can use this program to compare the common and different files between two folders.
	Author: Zhengdeng Lei (zlei2@uic.edu)

	'''
	epi_log='''
This program requires the python packages:
1. argparse
2. openpyxl (pypm install "openpyxl<=2.1.5")
3. hashlib
	'''	
	help_help = '''\
	show this help message and exit\
	'''
	version_help = '''\
	show the version of this program\
	'''
	#command_file_help = '''\
	#pipeline command file (e.g. pipeline_commands.sh).
	#Steps are separated by an empty line. 
	#Command lines within each step will be run simultaneously.
	#'''

	a_dir_help = '''\
	directory A (with DOUBLE QUOTE)
	'''
	b_dir_help = '''\
	directory B (with DOUBLE QUOTE)
	'''
	excel_out_help = '''\
	prefix of the output Excel file which contains the file comparison results.
	'''
	


	arg_parser = argparse.ArgumentParser(description=main_description, epilog=epi_log, formatter_class=argparse.RawTextHelpFormatter, add_help=False)

	###############################
	#    1. required arguments    #
	###############################
	required_group = arg_parser.add_argument_group("required arguments")
	required_group.add_argument("-a", dest="a_dir", action="store", required=True, default=None, help=a_dir_help)
	required_group.add_argument("-b", dest="b_dir", action="store", required=True, default=None, help=b_dir_help)



	###############################
	#    2. optional arguments    #
	###############################
	optional_group = arg_parser.add_argument_group("optional arguments")
	optional_group.add_argument("-p", dest="prefix", metavar="prefix of the output Excel file", default=None, help=excel_out_help)
	###############################
	#    3. positional arguments  #
	###############################
	#arg_parser.add_argument('command_file', metavar="COMMAND_FILE", type=str, help=command_file_help)


	optional_group.add_argument("-h", "--help", action="help", help=help_help)
	optional_group.add_argument("-v", "--version", action="version", version="%(prog)s: version 1.0", help=version_help)

	args = arg_parser.parse_args()
	args.a_dir = os.path.abspath(args.a_dir)
	args.b_dir = os.path.abspath(args.b_dir)
	
	crrt_time = datetime.datetime.now()
	crrt_time = "%4d-%02d-%02d_%02d-%02d-%02d" % (crrt_time.year, crrt_time.month, crrt_time.day, crrt_time.hour, crrt_time.minute, crrt_time.second)
	filename  = "compare_dir_" + crrt_time + ".xlsx"
	if args.prefix:
		args.excel_out = os.path.join(os.getcwd(), args.prefix + "_compare_dir_" + crrt_time + ".xlsx")
	else:
		args.excel_out = os.path.join(os.getcwd(), "compare_dir_" + crrt_time + ".xlsx")
		


	create_dirs = []
	for create_dir in create_dirs:
		dir_path = eval("args." + create_dir)
		if not os.path.exists(dir_path):
			print dir_path, "is created!\n"
			os.makedirs(dir_path)

	must_exist_paths = ["a_dir", "b_dir"]
	for must_exist_path in must_exist_paths:
		path = eval("args." + must_exist_path)
		if not os.path.exists(path):
			arg_parser.print_help()
			print "\n\nError: " + path + " does not exist!\n"
			sys.exit()

	#if not os.path.dirname(args.excel_out):
	#	args.excel_out = os.path.join(os.getcwd(),args.excel_out)
	return args
	#args = get_arguments()
	#return [args, arg_parser]
	#[args, arg_parser] = get_arguments()


	 
if __name__ == '__main__':
	main()



